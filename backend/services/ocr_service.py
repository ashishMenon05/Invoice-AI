import io
import logging
from PIL import Image
import pdfplumber
import pytesseract
import cv2
import numpy as np

logger = logging.getLogger(__name__)

def extract_text_from_file(file_bytes: bytes, filename: str) -> str:
    """
    Extracts readable text from an uploaded invoice file.
    - PDF: pdfplumber (digital) → pytesseract (scanned fallback)
    - Image (jpg/png): pytesseract
    - Excel (xlsx/xls): openpyxl — each sheet rendered as tabular text
    - CSV: decoded directly as text
    """
    ext = filename.lower().split('.')[-1]
    
    if ext == "pdf":
        return _extract_from_pdf(file_bytes)
    elif ext in ["png", "jpg", "jpeg"]:
        return _extract_from_image(file_bytes)
    elif ext in ["xlsx", "xls"]:
        return _extract_from_excel(file_bytes)
    elif ext == "csv":
        return _extract_from_csv(file_bytes)
    else:
        logger.warning(f"Unsupported file format for OCR: {ext}")
        return ""

def _extract_from_pdf(file_bytes: bytes) -> str:
    extracted_text = ""
    try:
        # First attempt: Digital text extraction
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    extracted_text += text + "\n"
        
        # If pure digital extraction is empty or too short, fallback to OCR on pages
        if len(extracted_text.strip()) < 50:
            logger.info("PDF appears scanned. Falling back to OCR.")
            extracted_text = ""
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    # Convert page to image
                    im = page.to_image(resolution=200) # Increased resolution for better OCR
                    pil_image = im.original.convert("RGB")
                    
                    # Preprocess with OpenCV
                    cv_img = np.array(pil_image)
                    cv_img = cv_img[:, :, ::-1].copy() # RGB to BGR
                    gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
                    
                    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
                    processed_image = Image.fromarray(thresh)
                    
                    custom_config = r'--oem 3 --psm 6'
                    text = pytesseract.image_to_string(processed_image, config=custom_config, timeout=30)
                    extracted_text += text + "\n"
    except Exception as e:
        logger.error(f"Error extracting from PDF: {e}")
        
    return extracted_text

def _extract_from_image(file_bytes: bytes) -> str:
    try:
        image = Image.open(io.BytesIO(file_bytes))
        # Ensure image is in a format pytesseract likes
        if image.mode != "RGB":
            image = image.convert("RGB")
            
        # Advanced OpenCV Preprocessing for Invoice read optimization
        cv_img = np.array(image)
        # Convert RGB to BGR for cv2
        cv_img = cv_img[:, :, ::-1].copy()
        
        # 1. Grayscale
        gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
        
        # 2. Adaptive Binarization (pulls dark text from light backgrounds seamlessly)
        thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
        
        processed_image = Image.fromarray(thresh)

        # Hard 30s timeout kills the subprocess if it hangs on complex noisy jpegs
        custom_config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(processed_image, config=custom_config, timeout=30)
        return text
    except RuntimeError as timeout_e:
        # Pytesseract throws RuntimeError on timeout. We re-raise as generic Exception to trip the AI fallback loop
        logger.error(f"OCR Timeout: {timeout_e}")
        raise Exception(f"OCR stalled: {timeout_e}")
    except Exception as e:
        logger.error(f"Error extracting from image: {e}")
        return ""

def _extract_from_excel(file_bytes: bytes) -> str:
    """
    Reads an Excel workbook and converts every sheet to tab-separated text
    so the LLM can parse it as a tabular invoice document.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                # Skip completely empty rows
                if any(cell is not None for cell in row):
                    parts.append("\t".join(str(c) if c is not None else "" for c in row))
        text = "\n".join(parts)
        logger.info(f"Excel extraction done — {len(text)} chars from {len(wb.sheetnames)} sheet(s).")
        return text
    except Exception as e:
        logger.error(f"Excel extraction failed: {e}")
        return ""

def _extract_from_csv(file_bytes: bytes) -> str:
    """Decodes CSV bytes as plain text for the LLM."""
    try:
        return file_bytes.decode("utf-8", errors="replace")
    except Exception as e:
        logger.error(f"CSV decode failed: {e}")
        return ""
